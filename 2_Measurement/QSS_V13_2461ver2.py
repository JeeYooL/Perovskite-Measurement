# ==============================================================================
#      Solar Cell Analyzer: v19.1 (Keithley 2461 Compatible)
#      Target: Keithley 2461 (SCPI Mode)
#      Modified for High Current Capability (up to 10A)
# ==============================================================================

import tkinter as tk
from tkinter import ttk, Button, Label, Entry, messagebox, filedialog, scrolledtext, Radiobutton, IntVar, BooleanVar
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import pyvisa
import numpy as np
import time
import os
import logging
import threading
import json
from datetime import datetime
from scipy.interpolate import interp1d
from scipy.stats import linregress 
import pandas as pd
from openpyxl import load_workbook, Workbook 
from openpyxl.styles import Font, PatternFill
import re

# ==============================================================================
# 1. Global Variables
# ==============================================================================
stop_flag, is_paused = False, False
thread = None
keithley = None
global_save_path = ""
instrument_info = "Unknown"

plots_data = {
    'J-V Scan':    {'x': [], 'y': [], 'ax_label': ("Voltage (V)", "Current Density (mA/cm²)"), 'color': 'b'},
    'QSS-IV':      {'x': [], 'y': [], 'ax_label': ("Voltage (V)", "Current Density (mA/cm²)"), 'color': 'r'},
    'SPO (PCE)':   {'x': [], 'y': [], 'ax_label': ("Time (s)", "Efficiency (%)"), 'color': 'r'},
    'Imp Track':   {'x': [], 'y': [], 'ax_label': ("Time (s)", "Current Density (mA/cm²)"), 'color': 'c'},
    'Vmp Track':   {'x': [], 'y': [], 'ax_label': ("Time (s)", "Voltage (V)"), 'color': 'g'},
    'Power Track': {'x': [], 'y': [], 'ax_label': ("Time (s)", "Power (mW/cm²)"), 'color': 'm'},
}

class TextHandler(logging.Handler):
    def __init__(self, text_widget):
        super().__init__(); self.text_widget = text_widget
    def emit(self, record):
        msg = self.format(record)
        def append():
            self.text_widget.configure(state='normal'); self.text_widget.insert(tk.END, msg + '\n')
            self.text_widget.configure(state='disabled'); self.text_widget.yview(tk.END)
        self.text_widget.after(0, append)

# ==============================================================================
# 2. GUI Setup (V13/V16 Layout)
# ==============================================================================
win = tk.Tk(); win.title("Solar Cell Analyzer for LKChem")
win.geometry("1650x1000")

# Smart Pixel Logic
pixel_var = IntVar(value=1)
hotkey_run_var = BooleanVar(value=False) # Safety Switch

def smart_pixel_trigger(event, num):
    current = pixel_var.get()
    if current == num:
        if hotkey_run_var.get():
            if btn_jv_single['state'] == tk.NORMAL: 
                logging.info(f"[Hotkey {num}] Starting JV Scan...")
                start_thread("JV")
        else:
            logging.info(f"[Hotkey {num}] Ignored (Enable 'Key-Run')")
    else:
        if hotkey_run_var.get():
            pixel_var.set(num)
            logging.info(f"[Hotkey {num}] Active Pixel Changed to: {num}")
        else:
            logging.info(f"[Hotkey {num}] Pixel change ignored (Enable 'Key-Run')")

win.bind('1', lambda e: smart_pixel_trigger(e, 1))
win.bind('2', lambda e: smart_pixel_trigger(e, 2))
win.bind('3', lambda e: smart_pixel_trigger(e, 3))
win.bind('4', lambda e: smart_pixel_trigger(e, 4))

left_panel = tk.Frame(win, padx=5, pady=5); left_panel.pack(side=tk.LEFT, fill=tk.Y, expand=False)
right_panel = tk.Frame(win, padx=5, pady=5); right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

param_entries = {} 

# --- SECTION 1: Common ---
common_frame = tk.LabelFrame(left_panel, text="1. Common Settings & Recipe", padx=5, pady=5, font=("Arial", 10, "bold"))
common_frame.pack(fill=tk.X, pady=(0, 5))

def select_global_folder():
    global global_save_path
    path = filedialog.askdirectory(title="Select Root Folder")
    if path: 
        today_folder = datetime.now().strftime("%Y%m%d")
        final_path = os.path.join(path, today_folder)
        os.makedirs(final_path, exist_ok=True)
        global_save_path = final_path
        lbl_path.config(text=f".../{today_folder}")
        entry_save_path.config(state='normal'); entry_save_path.delete(0, tk.END); entry_save_path.insert(0, today_folder); entry_save_path.config(state='readonly')

def save_preset():
    f = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON", "*.json")], title="Save Preset")
    if f:
        data = {k: v.get() for k, v in param_entries.items()}
        data['auto_vmp'] = auto_vmp_var.get()
        with open(f, 'w') as json_file: json.dump(data, json_file, indent=4)
        logging.info(f"Preset saved: {os.path.basename(f)}")
def load_preset():
    f = filedialog.askopenfilename(filetypes=[("JSON", "*.json")], title="Load Preset")
    if f:
        try:
            with open(f, 'r') as json_file: data = json.load(json_file)
            for k, v in data.items():
                if k == 'auto_vmp': auto_vmp_var.set(v); continue
                if k in param_entries:
                    if isinstance(param_entries[k], ttk.Combobox): param_entries[k].set(v)
                    else: param_entries[k].delete(0, tk.END); param_entries[k].insert(0, v)
            logging.info(f"Preset loaded: {os.path.basename(f)}")
        except Exception as e: messagebox.showerror("Error", f"Load failed: {e}")

btn_path = Button(common_frame, text="📂 Select Root", command=select_global_folder, bg="#fff9c4", width=12)
btn_path.grid(row=0, column=0, padx=2, pady=2)
lbl_path = Label(common_frame, text="No folder", fg="gray", font=("Arial", 8))
lbl_path.grid(row=0, column=1, sticky='w')
Button(common_frame, text="Save Preset", command=save_preset, bg="#e0f7fa", width=10).grid(row=0, column=2, padx=2)
Button(common_frame, text="Load Preset", command=load_preset, bg="#e0f7fa", width=10).grid(row=0, column=3, padx=2)

tk.Label(common_frame, text="User:").grid(row=1, column=0, sticky='e')
e_user = Entry(common_frame, width=12); e_user.insert(0, "User"); e_user.grid(row=1, column=1, sticky='w')
param_entries["User Name"] = e_user
tk.Label(common_frame, text="Saved Folder:").grid(row=1, column=2, sticky='e')
entry_save_path = Entry(common_frame, width=15, state='readonly'); entry_save_path.grid(row=1, column=3, sticky='w')

tk.Label(common_frame, text="Device ID:").grid(row=2, column=0, sticky='e')
e_dev = Entry(common_frame, width=40); e_dev.insert(0, "Perovskite_Cell"); e_dev.grid(row=2, column=1, columnspan=3, sticky='w', padx=2, pady=2)
param_entries["Device ID"] = e_dev

pixel_frame = tk.Frame(common_frame)
pixel_frame.grid(row=3, column=0, columnspan=4, sticky='w', pady=2)
tk.Label(pixel_frame, text="Active Pixel:", font=("Arial", 9, "bold"), fg="darkgreen").pack(side=tk.LEFT, padx=5)
for i in range(1, 5): Radiobutton(pixel_frame, text=f"{i}", variable=pixel_var, value=i).pack(side=tk.LEFT)
tk.Label(pixel_frame, text=" | ", fg="gray").pack(side=tk.LEFT, padx=5)
tk.Checkbutton(pixel_frame, text="Key-Run", variable=hotkey_run_var, fg="red", font=("Arial", 9)).pack(side=tk.LEFT)

def scan_ports():
    try: rm = pyvisa.ResourceManager(); gpib_combo['values'] = rm.list_resources(); gpib_combo.current(0) if gpib_combo['values'] else None; logging.info("Ports scanned.")
    except: logging.error("No ports found.")
def check_conn():
    try: 
        rm = pyvisa.ResourceManager(); inst = rm.open_resource(gpib_combo.get())
        inst.write_termination='\n'; inst.read_termination='\n'
        global instrument_info; instrument_info = inst.query("*IDN?").strip()
        messagebox.showinfo("Connected", instrument_info); inst.close()
    except Exception as e: messagebox.showerror("Error", str(e))

tk.Label(common_frame, text="Area (cm²)").grid(row=4, column=0, sticky='e')
e_area = Entry(common_frame, width=8); e_area.insert(0, "0.096"); e_area.grid(row=4, column=1, sticky='w')
param_entries["Active Area (cm²)"] = e_area

tk.Label(common_frame, text="GPIB").grid(row=4, column=2, sticky='e')
gpib_combo = ttk.Combobox(common_frame, width=12); gpib_combo.grid(row=4, column=3, sticky='w')
param_entries["GPIB Address"] = gpib_combo
Button(common_frame, text="Scan", command=scan_ports, width=5).grid(row=4, column=4)
Button(common_frame, text="Conn", command=check_conn, width=5).grid(row=4, column=5)

# [MODIFIED] Added 2461 High Current Options
tk.Label(common_frame, text="I-Limit").grid(row=5, column=0, sticky='e')
ilim_combo = ttk.Combobox(common_frame, values=["10 mA", "100 mA", "1 A", "4 A", "5 A", "10 A"], width=8, state="readonly"); ilim_combo.set("1 A"); ilim_combo.grid(row=5, column=1, sticky='w')
param_entries["Current Limit"] = ilim_combo
tk.Label(common_frame, text="Sense").grid(row=5, column=2, sticky='e')
sense_combo = ttk.Combobox(common_frame, values=["2-Wire (Local)", "4-Wire (Remote)"], width=12, state="readonly"); sense_combo.set("4-Wire (Remote)"); sense_combo.grid(row=5, column=3, sticky='w')
param_entries["Sense Mode"] = sense_combo


# --- SECTION 2: Results ---
result_main_frame = tk.LabelFrame(left_panel, text="Measurement Results", padx=2, pady=2, font=("Arial", 11, "bold"), fg="blue")
result_main_frame.pack(fill=tk.X, pady=(0, 5))

jv_res_frame = tk.LabelFrame(result_main_frame, text="Fast J-V / Hysteresis", padx=2, pady=2, fg="blue")
jv_res_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
qss_res_frame = tk.LabelFrame(result_main_frame, text="QSS-IV (Std)", padx=2, pady=2, fg="green")
qss_res_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
nrel_res_frame = tk.LabelFrame(result_main_frame, text="QSS-NREL", padx=2, pady=2, fg="purple")
nrel_res_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)

def create_result_box(parent, label_text, row, col, color="yellow"):
    frame = tk.Frame(parent, padx=1, pady=1)
    frame.grid(row=row, column=col, sticky="ew", padx=2, pady=1)
    tk.Label(frame, text=label_text, font=("Arial", 8)).pack(anchor="w")
    lbl_val = tk.Label(frame, text="0.000", font=("Arial", 11, "bold"), bg=color, width=7, relief="sunken")
    lbl_val.pack(fill=tk.X)
    return lbl_val

res_labels = {"JV": {}, "QSS": {}, "NREL": {}}
frames = [jv_res_frame, qss_res_frame, nrel_res_frame]
keys = ["JV", "QSS", "NREL"]
for i, key in enumerate(keys):
    res_labels[key]['Jsc'] = create_result_box(frames[i], "Jsc", 0, 0)
    res_labels[key]['Voc'] = create_result_box(frames[i], "Voc", 0, 1)
    res_labels[key]['FF']  = create_result_box(frames[i], "FF", 1, 0)
    res_labels[key]['PCE'] = create_result_box(frames[i], "PCE", 1, 1, color="#ADFF2F")

res_labels['JV']['Rsh'] = create_result_box(jv_res_frame, "Rsh", 0, 2, color="lightcyan")
res_labels['JV']['Rs']  = create_result_box(jv_res_frame, "Rs", 1, 2, color="lightcyan")
res_labels['JV']['HI']  = create_result_box(jv_res_frame, "H-Index", 0, 3, color="orange")
jv_labels = res_labels['JV']; qss_labels = res_labels['QSS']

# --- SECTION 3: Settings ---
settings_container = tk.Frame(left_panel)
settings_container.pack(fill=tk.BOTH, expand=True, pady=5)
for i in range(3): settings_container.grid_columnconfigure(i, weight=1)

def add_entry_grid(parent, label, default, row):
    tk.Label(parent, text=label).grid(row=row, column=0, sticky='w', pady=1)
    e = tk.Entry(parent, width=9); e.insert(0, str(default)); e.grid(row=row, column=1, sticky='e', padx=1)
    param_entries[label] = e

# [2. JV]
jv_frame = tk.LabelFrame(settings_container, text="2. Fast J-V Sweep", padx=5, pady=5, fg="blue", font=("Arial", 10, "bold"))
jv_frame.grid(row=0, column=0, sticky="nsew", padx=2, pady=2)
def update_jv_dir(event):
    mode = param_entries["JV Direction"].get(); s = param_entries["JV Start V"].get(); e = param_entries["JV End V"].get()
    try: s_val, e_val = float(s), float(e)
    except: return
    if mode == "Reverse (Hi->Lo)":
        param_entries["JV Start V"].delete(0, tk.END); param_entries["JV Start V"].insert(0, max(s_val, e_val))
        param_entries["JV End V"].delete(0, tk.END); param_entries["JV End V"].insert(0, min(s_val, e_val))
    else:
        param_entries["JV Start V"].delete(0, tk.END); param_entries["JV Start V"].insert(0, min(s_val, e_val))
        param_entries["JV End V"].delete(0, tk.END); param_entries["JV End V"].insert(0, max(s_val, e_val))
tk.Label(jv_frame, text="Direction").grid(row=0, column=0, sticky='w')
jv_dir_combo = ttk.Combobox(jv_frame, values=["Reverse (Hi->Lo)", "Forward (Lo->Hi)"], width=10, state="readonly"); jv_dir_combo.set("Reverse (Hi->Lo)")
jv_dir_combo.grid(row=0, column=1, sticky='e'); jv_dir_combo.bind("<<ComboboxSelected>>", update_jv_dir); param_entries["JV Direction"] = jv_dir_combo
add_entry_grid(jv_frame, "JV Start V", "1.25", 1); add_entry_grid(jv_frame, "JV End V", "-0.05", 2)
add_entry_grid(jv_frame, "JV Step (mV)", "20", 3); add_entry_grid(jv_frame, "JV Delay (ms)", "50", 4)

# [3. QSS]
qss_frame = tk.LabelFrame(settings_container, text="3. QSS-IV (Std)", padx=5, pady=5, fg="green", font=("Arial", 10, "bold"))
qss_frame.grid(row=0, column=1, sticky="nsew", padx=2, pady=2)
def update_qss_dir(event):
    mode = param_entries["QSS Direction"].get(); s = param_entries["QSS Start V"].get(); e = param_entries["QSS End V"].get()
    try: s_val, e_val = float(s), float(e)
    except: return
    if mode == "Reverse (Hi->Lo)":
        param_entries["QSS Start V"].delete(0, tk.END); param_entries["QSS Start V"].insert(0, max(s_val, e_val))
        param_entries["QSS End V"].delete(0, tk.END); param_entries["QSS End V"].insert(0, min(s_val, e_val))
    else:
        param_entries["QSS Start V"].delete(0, tk.END); param_entries["QSS Start V"].insert(0, min(s_val, e_val))
        param_entries["QSS End V"].delete(0, tk.END); param_entries["QSS End V"].insert(0, max(s_val, e_val))
tk.Label(qss_frame, text="Direction").grid(row=0, column=0, sticky='w')
qss_dir_combo = ttk.Combobox(qss_frame, values=["Reverse (Hi->Lo)", "Forward (Lo->Hi)"], width=10, state="readonly"); qss_dir_combo.set("Forward (Lo->Hi)")
qss_dir_combo.grid(row=0, column=1, sticky='e'); qss_dir_combo.bind("<<ComboboxSelected>>", update_qss_dir); param_entries["QSS Direction"] = qss_dir_combo
add_entry_grid(qss_frame, "QSS Start V", "-0.05", 1); add_entry_grid(qss_frame, "QSS End V", "1.25", 2)
add_entry_grid(qss_frame, "QSS Step (mV)", "65", 3); add_entry_grid(qss_frame, "Delay (Init)", "100", 4)

# [4. MPPT]
mppt_frame = tk.LabelFrame(settings_container, text="4. MPPT Settings", padx=5, pady=5, fg="red", font=("Arial", 10, "bold"))
mppt_frame.grid(row=0, column=2, sticky="nsew", padx=2, pady=2)
add_entry_grid(mppt_frame, "MPPT Start V (V)", "0.9", 0); add_entry_grid(mppt_frame, "Perturb Step (mV)", "10", 1)
add_entry_grid(mppt_frame, "MPPT Int (s)", "0.5", 2); add_entry_grid(mppt_frame, "Duration (h)", "1.0", 3)
auto_vmp_var = BooleanVar(value=True)
# [New] Time Unit Selection
tk.Label(mppt_frame, text="X-Axis Unit").grid(row=4, column=0, sticky='w')
mppt_unit_combo = ttk.Combobox(mppt_frame, values=["sec", "min", "hour", "day"], width=8, state="readonly")
mppt_unit_combo.set("sec")
mppt_unit_combo.grid(row=4, column=1, sticky='w')
param_entries["MPPT Unit"] = mppt_unit_combo

tk.Checkbutton(mppt_frame, text="Auto-Find Vmp", variable=auto_vmp_var, fg="darkred").grid(row=5, column=0, columnspan=2, sticky='w')

# [5. Loop]
loop_frame = tk.LabelFrame(settings_container, text="5. Loop Settings", padx=5, pady=5, fg="darkblue", font=("Arial", 10, "bold"))
loop_frame.grid(row=1, column=0, sticky="nsew", padx=2, pady=2)
add_entry_grid(loop_frame, "Repeat Count", "5", 0); add_entry_grid(loop_frame, "Loop Int (s)", "60", 1)

# [6. NREL]
nrel_frame = tk.LabelFrame(settings_container, text="6. QSS-NREL Settings", padx=5, pady=5, fg="purple", font=("Arial", 10, "bold"))
nrel_frame.grid(row=1, column=1, sticky="nsew", padx=2, pady=2)
add_entry_grid(nrel_frame, "Target Points", "15", 0); add_entry_grid(nrel_frame, "Range (±V)", "0.15", 1)
add_entry_grid(nrel_frame, "NREL Wait (s)", "60.0", 2); add_entry_grid(nrel_frame, "Threshold (%)", "0.07", 3)
add_entry_grid(nrel_frame, "Max Wait (s)", "15.0", 4)

# [7. SPO]
spo_frame = tk.LabelFrame(settings_container, text="7. SPO (Stabilized Power)", padx=5, pady=5, fg="darkorange", font=("Arial", 10, "bold"))
spo_frame.grid(row=1, column=2, sticky="nsew", padx=2, pady=2)
add_entry_grid(spo_frame, "SPO Voltage (V)", "0.95", 0)
add_entry_grid(spo_frame, "SPO Duration (s)", "300", 1)
add_entry_grid(spo_frame, "SPO Interval (s)", "1.0", 2)
Label(spo_frame, text="Fixed V Tracking", font=("Arial", 8), fg="gray").grid(row=3, column=0, columnspan=2, pady=(5,0))

# --- Control Buttons ---
action_frame = tk.LabelFrame(left_panel, text="Experiment Control", padx=5, pady=5)
action_frame.pack(fill=tk.X, pady=5)
action_frame.grid_columnconfigure(0, weight=1); action_frame.grid_columnconfigure(1, weight=1); action_frame.grid_columnconfigure(2, weight=1)

Label(action_frame, text="Single Run", font=("Arial", 9, "bold")).grid(row=0, column=0, columnspan=3, sticky="w", padx=5)
btn_jv_single = Button(action_frame, text="JV Scan", command=lambda: start_thread("JV"), bg="#e3f2fd", height=2)
btn_jv_single.grid(row=1, column=0, sticky="ew", padx=2, pady=2)
btn_qss_single = Button(action_frame, text="QSS-IV", command=lambda: start_thread("QSS"), bg="#f3e5f5", height=2)
btn_qss_single.grid(row=1, column=1, sticky="ew", padx=2, pady=2)
btn_mppt_single = Button(action_frame, text="MPPT", command=lambda: start_thread("MPPT"), bg="#ffebee", height=2)
btn_mppt_single.grid(row=1, column=2, sticky="ew", padx=2, pady=2)

Label(action_frame, text="Loop Run", font=("Arial", 9, "bold")).grid(row=2, column=0, columnspan=3, sticky="w", padx=5, pady=(5,0))
btn_jv_loop = Button(action_frame, text="Loop JV", command=lambda: start_thread("Loop_JV"), bg="#bbdefb", height=2)
btn_jv_loop.grid(row=3, column=0, sticky="ew", padx=2, pady=2)
btn_qss_loop = Button(action_frame, text="Loop QSS", command=lambda: start_thread("Loop_QSS"), bg="#e1bee7", height=2)
btn_qss_loop.grid(row=3, column=1, sticky="ew", padx=2, pady=2)
btn_hyst_loop = Button(action_frame, text="Loop Hyst", command=lambda: start_thread("Loop_Hysteresis"), bg="#c5cae9", height=2)
btn_hyst_loop.grid(row=3, column=2, sticky="ew", padx=2, pady=2)

Label(action_frame, text="Special", font=("Arial", 9, "bold")).grid(row=4, column=0, columnspan=3, sticky="w", padx=5, pady=(5,0))
btn_hyst_single = Button(action_frame, text="Hysteresis", command=lambda: start_thread("Hysteresis"), bg="#e8eaf6", height=2)
btn_hyst_single.grid(row=5, column=0, sticky="ew", padx=2, pady=2)
btn_qss_nrel = Button(action_frame, text="QSS-NREL", command=lambda: start_thread("QSS-NREL"), bg="#d1c4e9", height=2, fg="darkblue")
btn_qss_nrel.grid(row=5, column=1, sticky="ew", padx=2, pady=2)
btn_spo = Button(action_frame, text="SPO", command=lambda: start_thread("SPO"), bg="#fff3e0", height=2, fg="darkorange")
btn_spo.grid(row=5, column=2, sticky="ew", padx=2, pady=2)

btn_stop = Button(action_frame, text="STOP ALL", command=lambda: stop_exp(), state=tk.DISABLED, bg="orange", height=2)
btn_stop.grid(row=6, column=0, columnspan=3, sticky="ew", padx=2, pady=5)

status_lbl = Label(left_panel, text="Ready", bg="gray", fg="white", font=("Arial", 12)); status_lbl.pack(fill=tk.X, pady=2)
log_txt = scrolledtext.ScrolledText(left_panel, height=8, font=("Courier", 8)); log_txt.pack(fill=tk.BOTH, expand=True)

# --- Graph ---
notebook = ttk.Notebook(right_panel); notebook.pack(fill=tk.BOTH, expand=True)
for name, p_data in plots_data.items():
    tab = ttk.Frame(notebook); notebook.add(tab, text=name)
    fig = Figure(figsize=(4, 3.5), dpi=100); ax = fig.add_subplot(111)
    p_data['fig'] = fig; p_data['ax'] = ax
    canvas = FigureCanvasTkAgg(fig, master=tab); canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    NavigationToolbar2Tk(canvas, tab).update()
    p_data['canvas'] = canvas

# ==============================================================================
# 3. Core Logic
# ==============================================================================
def setup_logger():
    logger = logging.getLogger(); logger.setLevel(logging.INFO); logger.handlers = []; logger.addHandler(TextHandler(log_txt))
def get_param(key): return param_entries[key].get()
def update_status(msg, color): status_lbl.config(text=msg, bg=color)

def init_keithley(rm, addr):
    k = rm.open_resource(addr); k.write_termination = '\n'; k.read_termination = '\n'; k.timeout = 25000; k.write('*RST')
    
    # [MODIFIED] Added 2461 High Current Ranges to Map
    limit_str = get_param("Current Limit")
    ilim_map = {
        "10 mA": 10e-3, "100 mA": 100e-3, "1 A": 1.0,
        "4 A": 4.0, "5 A": 5.0, "10 A": 10.0
    }
    ilim = ilim_map.get(limit_str, 1.0)
    
    k.write(':SOUR:FUNC VOLT')
    # 2461 supports same syntax for standard DC operation
    k.write(f':SOUR:VOLT:ILIM {ilim}') 
    k.write(':SENS:FUNC "CURR"'); k.write(':SENS:CURR:RANG:AUTO ON')
    k.write(':SENS:CURR:NPLC 1') 
    
    if "4-Wire" in get_param("Sense Mode"): 
        k.write(':SENS:VOLT:RSEN ON'); k.write(':SENS:CURR:RSEN ON')
    else: 
        k.write(':SENS:VOLT:RSEN OFF'); k.write(':SENS:CURR:RSEN OFF')
    return k

def check_compliance(keithley, ilim, v_set):
    # [MODIFIED] Added 2461 High Current Ranges to Map
    ilim_map = {
        "10 mA": 10e-3, "100 mA": 100e-3, "1 A": 1.0,
        "4 A": 4.0, "5 A": 5.0, "10 A": 10.0
    }
    limit_val = ilim_map.get(ilim, 1.0)
    try:
        curr = float(keithley.query(':MEAS:CURR?'))
        if abs(curr) >= limit_val * 0.98: logging.warning(f"⚠️ COMPLIANCE REACHED! ({curr:.3f} A)")
        return curr
    except: return 0.0

def measure_current_with_averaging(keithley, n_avg=5, delay_per_point=0.01):
    measurements = []
    for _ in range(n_avg):
        curr = float(keithley.query(':MEAS:CURR?')); measurements.append(curr); time.sleep(delay_per_point)
    return np.median(measurements), np.std(measurements)

def is_qss_stabilized_advanced(history, threshold=0.05):
    if len(history) < 20: return False, None
    recent = history[-10:]; mean_val = np.mean(recent)
    if abs(mean_val) < 1e-9: return False, None
    std_dev = np.std(recent) / abs(mean_val) * 100
    x_vals = np.arange(len(recent)); slope, _, _, _, _ = linregress(x_vals, recent)
    normalized_slope = abs(slope / mean_val) * 100
    range_var = (np.max(recent) - np.min(recent)) / abs(mean_val) * 100
    is_stable = (std_dev < threshold) and (normalized_slope < 0.01) and (range_var < threshold * 2)
    metrics = {'std_dev': std_dev, 'slope': normalized_slope, 'range': range_var}
    return is_stable, metrics

def calculate_jv_params_enhanced(v_arr, j_arr):
    try:
        if len(v_arr) < 5: return 0, 0, 0, 0, 0, 0, 0
        sorted_indices = np.argsort(v_arr); v_sorted = np.array(v_arr)[sorted_indices]; j_sorted = np.array(j_arr)[sorted_indices]
        jsc = float(np.interp(0, v_sorted, j_sorted)) if v_sorted.min() <= 0 <= v_sorted.max() else 0.0
        voc = 0.0
        if j_sorted.min() <= 0 <= j_sorted.max(): voc = float(interp1d(j_sorted, v_sorted, kind='linear', fill_value='extrapolate')(0))
        p_arr = v_sorted * j_sorted; pmax = np.max(p_arr); vmp = v_sorted[np.argmax(p_arr)]
        ff = (pmax / (jsc * voc)) * 100 if jsc > 0 and voc > 0 else 0.0
        r_sh, r_s = 0.0, 0.0
        mask_sh = (v_sorted > -0.05) & (v_sorted < 0.05)
        if np.sum(mask_sh) > 2:
            slope, _, _, _, _ = linregress(v_sorted[mask_sh], j_sorted[mask_sh])
            r_sh = abs(1/slope) * 1000 if slope != 0 else 0
        mask_s = (v_sorted > voc - 0.1) & (v_sorted < voc + 0.1)
        if np.sum(mask_s) > 2:
             slope, _, _, _, _ = linregress(v_sorted[mask_s], j_sorted[mask_s])
             r_s = abs(1/slope) * 1000 if slope != 0 else 0
        return jsc, voc, ff, pmax, vmp, r_sh, r_s
    except: return 0, 0, 0, 0, 0, 0, 0

def update_results_gui(mode, jsc, voc, ff, pce, hi=None, rsh=None, rs=None):
    if mode == "JV": target = res_labels['JV']
    elif mode == "QSS": target = res_labels['QSS']
    elif mode == "QSS_NREL": target = res_labels['NREL']
    else: return
    target['Jsc'].config(text=f"{jsc:.3f}"); target['Voc'].config(text=f"{voc:.3f}")
    target['FF'].config(text=f"{ff:.2f}"); target['PCE'].config(text=f"{pce:.2f}")
    if hi is not None and mode == "JV": res_labels['JV']['HI'].config(text=f"{hi:.2f}")
    if rsh is not None and 'Rsh' in target: target['Rsh'].config(text=f"{rsh:.1f}")
    if rs is not None and 'Rs' in target: target['Rs'].config(text=f"{rs:.1f}")

def clean_filename(name): return re.sub(r'[\\/*?:"<>|]', "_", str(name)).strip()

def save_formatted_excel(mode, headers, data_rows, context, summary_stats=None, suffix=""):
    if not global_save_path: logging.error("Save failed: No global path."); return
    dev_id = clean_filename(context['dev_id']); pixel_num = context['pixel']
    date_str = datetime.now().strftime("%Y-%m-%d"); timestamp = datetime.now().strftime("%y%m%d_%H%M%S")
    filename = f"{dev_id}_{date_str}.xlsx"; filepath = os.path.join(global_save_path, filename)
    sheet_name = f"{mode}_P{pixel_num}"
    df_new = pd.DataFrame(data_rows, columns=headers)
    sample_header = f"{dev_id}-P{pixel_num}-{timestamp}{suffix}"
    
    if mode == "MPPT":
        df_new.rename(columns={"Elapsed(s)": "Time(s)", "Voltage(V)": "Voltage(V)", 
                               "Current(A)": sample_header, "Power(mW/cm2)": "P(mW/cm2)", "PCE(%)": "Eff(%)"}, inplace=True)
    else: 
        keep_cols = ["Voltage(V)", "Current(A)", "J(mA/cm2)"] 
        df_new = df_new[[c for c in keep_cols if c in df_new.columns]]
        df_new.rename(columns={"Voltage(V)": "Voltage(V)", "Current(A)": sample_header, "J(mA/cm2)": "J(mA/cm2)"}, inplace=True)

    meta_dict = {"User": context['user'], "Sample": f"{dev_id}-Px{pixel_num}", "Area": context['area'], 
                 "Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Mode": mode+suffix, "Instrument": instrument_info}
    
    if "JV" in mode:
        meta_dict.update({"Start V": context['jv_start'], "End V": context['jv_end'], "Scan Speed": f"{context['jv_step']}mV/{context['jv_delay']}ms"})
    elif "QSS" in mode:
        meta_dict.update({"Start V": context['qss_start'], "End V": context['qss_end'], "Wait": f"{context['max_wait']}s"})
    elif "MPPT" in mode:
        meta_dict.update({"Start V": context['mppt_start'], "Interval": f"{context['mppt_int']}s"})

    try:
        if not os.path.exists(filepath): wb = Workbook(); ws = wb.active; ws.title = "Summary_Cover"; wb.save(filepath)
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames: wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        start_col = 1 if ws.max_column == 1 and ws.cell(1, 1).value is None else ws.max_column + 2 
        row_idx = 1
        for k, v in meta_dict.items():
            ws.cell(row=row_idx, column=start_col, value=k).font = Font(bold=True); ws.cell(row=row_idx, column=start_col+1, value=v); row_idx += 1
        data_start_row = row_idx + 2
        for c_idx, col_name in enumerate(df_new.columns):
            cell = ws.cell(row=data_start_row, column=start_col + c_idx, value=col_name); cell.font = Font(bold=True); cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for r_idx, row_data in enumerate(df_new.values):
            for c_idx, val in enumerate(row_data): ws.cell(row=data_start_row + 1 + r_idx, column=start_col + c_idx, value=val)
        
        if mode in ["JV", "QSS", "QSS_NREL"] and summary_stats:
            last_data_row = data_start_row + len(df_new) + 2
            jsc, voc, ff, pce, _, rsh, rs = summary_stats
            stats_order = [("Jsc", jsc), ("Voc", voc), ("FF", ff), ("PCE (%)", pce), ("Rsh", rsh), ("Rs", rs)]
            ws.cell(row=last_data_row, column=start_col, value="Results").font = Font(bold=True)
            for i, (name, val) in enumerate(stats_order):
                r = last_data_row + 1 + i
                ws.cell(row=r, column=start_col, value=name); ws.cell(row=r, column=start_col+1, value=round(val, 2))
            
            if "Total_Summary" not in wb.sheetnames:
                sum_ws = wb.create_sheet("Total_Summary", 0)
                sum_ws.append(["Time", "Sample ID", "Pixel", "Mode", "Jsc", "Voc", "FF", "PCE", "H-Index", "Rsh", "Rs"])
            sum_ws = wb["Total_Summary"]
            hi_val = summary_stats[7] if len(summary_stats) > 7 else "-" 
            sum_ws.append([datetime.now().strftime("%H:%M:%S"), dev_id, pixel_num, mode+suffix, 
                           round(jsc,3), round(voc,3), round(ff,2), round(pce,2), str(hi_val), round(rsh,1), round(rs,1)])

        wb.save(filepath); logging.info(f"Saved Px{pixel_num} data to {filename}"); update_status(f"Saved Px{pixel_num}", "green")
    except PermissionError: messagebox.showerror("Save Error", f"Cannot save to '{filename}'.\nThe file is OPEN in Excel.")
    except Exception as e: logging.error(f"Save Failed: {e}"); messagebox.showerror("Save Error", str(e))

def run_jv_logic(context): run_sweep_logic(context, "JV")
def run_qss_logic(context): run_sweep_logic(context, "QSS")

def run_qss_nrel_logic(context):
    global keithley
    try:
        # 1. Pre-scan (Broad)
        logging.info("QSS-NREL: Fast Pre-scan...")
        s, e = float(context['jv_start']), float(context['jv_end'])
        _, vmp, _, voc, _, _, _ = run_sweep_internal(context, "JV", s, e, suffix="_PreCheck")
        
        if vmp == 0: logging.error("Pre-scan failed."); return

        # 2. Target Generation (I0, V0, Vmp cluster)
        logging.info(f"Vmp {vmp:.3f}V, Voc {voc:.3f}V. Targets set.")
        v_targets = np.linspace(vmp - 0.15, vmp + 0.15, int(get_param("Target Points")))
        v_targets = np.append(v_targets, [0.0, voc]) 
        v_targets = np.unique(np.sort(v_targets))
        v_targets = v_targets[(v_targets >= -0.2) & (v_targets <= 1.5)]
        
        s_qss, e_qss = float(context['qss_start']), float(context['qss_end'])
        if s_qss > e_qss: 
             v_targets = v_targets[::-1] 

        update_status("QSS-NREL Measuring...", "purple")
        data_captured = []; headers = ["Voltage(V)", "Current(A)", "J(mA/cm2)", "P(mW/cm2)", "Stabilized_Time(s)"]
        v_data, j_data = [] , []
        
        try: keithley.query('*IDN?') 
        except: rm = pyvisa.ResourceManager(); keithley = init_keithley(rm, context['gpib']); keithley.write(':OUTP ON')
        notebook.select(2) # NREL Tab
        
        max_wait = float(get_param("NREL Wait (s)")); threshold = float(context['threshold']); area = float(context['area'])

        for v_set in v_targets:
            if stop_flag: break
            keithley.write(f':SOUR:VOLT {v_set:.4f}')
            history = []; start_monitor = time.time(); is_stable = False; final_curr = 0.0
            while (time.time() - start_monitor) < max_wait:
                curr, _ = measure_current_with_averaging(keithley, n_avg=3)
                history.append(curr)
                if len(history) >= 10:
                    is_stable, _ = is_qss_stabilized_advanced(history, threshold)
                    if is_stable: final_curr = np.mean(history[-5:]); break
                time.sleep(0.5)
            if not is_stable: final_curr = history[-1]
            stab_time = time.time() - start_monitor
            j_val = -1 * (final_curr / area) * 1000; p_mW = v_set * j_val
            v_data.append(v_set); j_data.append(j_val)
            data_captured.append([v_set, final_curr, j_val, 0, stab_time])
            if len(v_data) % 1 == 0: win.after(0, update_plot_sweep, 'QSS-IV', v_data, j_data)

        win.after(0, update_plot_sweep, 'QSS-IV', v_data, j_data)
        jsc, voc, ff, pce, _, rsh, rs = calculate_jv_params_enhanced(v_data, j_data)
        win.after(0, update_results_gui, "QSS_NREL", jsc, voc, ff, pce, None, rsh, rs)
        if data_captured: save_formatted_excel("QSS_NREL", headers, data_captured, context, (jsc, voc, ff, pce, 0, rsh, rs))
            
    except Exception as e: logging.error(f"QSS-NREL Error: {e}")
    finally:
        if keithley: 
            try: keithley.write(':OUTP OFF'); keithley.close() 
            except: pass
        stop_exp()

def run_hyst_logic(context):
    logging.info("Starting Reverse Scan...")
    s, e = float(context['jv_start']), float(context['jv_end'])
    rev_start, rev_end = max(s, e), min(s, e); fwd_start, fwd_end = min(s, e), max(s, e)
    pce_rev, _, _, _, _, _, _ = run_sweep_internal(context, "JV", rev_start, rev_end, suffix="_Rev")
    time.sleep(1.0)
    logging.info("Starting Forward Scan...")
    pce_fwd, _, _, _, _, _, _ = run_sweep_internal(context, "JV", fwd_start, fwd_end, suffix="_Fwd")
    hi = 0.0
    if pce_rev > 0: hi = abs(pce_rev - pce_fwd) / pce_rev * 100 
    logging.info(f"Hysteresis Index: {hi:.2f}%")
    win.after(0, lambda: res_labels['JV']['HI'].config(text=f"{hi:.2f}"))

def run_sweep_internal(context, mode, start_v, end_v, suffix=""):
    global keithley, stop_flag
    data_captured = []; headers = ["Voltage(V)", "Current(A)", "J(mA/cm2)", "P(mW/cm2)"]
    if mode == "QSS": headers.append("Stabilized_Time(s)")
    v_data, j_data = [], [] 
    try:
        step_mv = float(context['jv_step']) if mode=="JV" else float(context['qss_step'])
        delay_ms = float(context['jv_delay']) if mode=="JV" else float(context['qss_delay'])
        max_wait = float(context['max_wait']); threshold = float(context['threshold'])
        area = float(context['area']); step_v = step_mv / 1000.0
        if start_v > end_v: step_v = -abs(step_v)
        else: step_v = abs(step_v)
        points = np.arange(start_v, end_v + (step_v/1000.0 if step_v > 0 else -step_v/1000.0), step_v)
        
        plot_key = 'J-V Scan' if mode == 'JV' else 'QSS-IV'
        notebook.select(0 if mode=="JV" else 1)
        try: keithley.query('*IDN?') 
        except: rm = pyvisa.ResourceManager(); keithley = init_keithley(rm, context['gpib']); keithley.write(':OUTP ON')

        # [AUTO-RANGE QSS LOGIC]
        if mode == "QSS" and context['auto_vmp']:
            logging.info("Auto-QSS: Running Pre-Scan to find Voc...")
            # Placeholder for potential pre-scan logic
            pass

        for v_set in points:
            if stop_flag: break
            keithley.write(f':SOUR:VOLT {v_set:.4f}')
            
            if mode == "QSS":
                # Smart QSS Logic
                history = []; start_monitor = time.time(); is_stable = False; final_curr = 0.0
                while (time.time() - start_monitor) < max_wait:
                    curr, _ = measure_current_with_averaging(keithley, n_avg=3)
                    history.append(curr)
                    if len(history) >= 10:
                        is_stable, _ = is_qss_stabilized_advanced(history, threshold)
                        if is_stable: final_curr = np.mean(history[-5:]); break
                    time.sleep(0.2)
                if not is_stable: final_curr = history[-1]
                stab_time = time.time() - start_monitor
                check_compliance(keithley, get_param("Current Limit"), v_set)
            else:
                # Multi-Average JV
                time.sleep(delay_ms / 1000.0)
                final_curr, _ = measure_current_with_averaging(keithley, n_avg=5, delay_per_point=0.005)
                check_compliance(keithley, get_param("Current Limit"), v_set)

            j_val = -1 * (final_curr / area) * 1000; v_data.append(v_set); j_data.append(j_val)
            row = [v_set, final_curr, j_val, 0] 
            if mode == "QSS": row.append(stab_time)
            data_captured.append(row)
            if len(v_data) % 2 == 0: win.after(0, update_plot_sweep, plot_key, v_data, j_data)

        win.after(0, update_plot_sweep, plot_key, v_data, j_data)
        jsc, voc, ff, pce, vmp, rsh, rs = calculate_jv_params_enhanced(v_data, j_data)
        win.after(0, update_results_gui, mode, jsc, voc, ff, pce, None, rsh, rs)
        if data_captured: save_formatted_excel(mode, headers, data_captured, context, (jsc, voc, ff, pce, vmp, rsh, rs), suffix=suffix)
        return pce, vmp, jsc, voc, ff, rsh, rs
    except Exception as e: logging.error(f"Error: {e}"); return 0.0, 0.0, 0, 0, 0, 0, 0
    finally:
        if suffix == "": 
            try: 
                keithley.write(':OUTP OFF')
                keithley.close()
            except: 
                pass

def run_sweep_logic(context, mode):
    start_v = float(context['jv_start']) if mode == "JV" else float(context['qss_start'])
    end_v = float(context['jv_end']) if mode == "JV" else float(context['qss_end'])
    try: run_sweep_internal(context, mode, start_v, end_v)
    finally: stop_exp()

def update_plot_sweep(key, x, y):
    ax = plots_data[key]['ax']; ax.cla(); col = plots_data[key]['color']
    ax.plot(x, y, color=col, marker='o', ms=3)
    ax.axhline(0, c='k', lw=0.5); ax.axvline(0, c='k', lw=0.5)
    ax.set_xlabel("Voltage (V)"); ax.set_ylabel("Current Density (mA/cm²)"); ax.set_title(key); ax.grid(True)
    plots_data[key]['canvas'].draw()

def run_mppt_logic(context):
    global keithley, stop_flag, is_paused
    data_captured = []; headers = ["Elapsed(s)", "Voltage(V)", "Current(A)", "Power(mW/cm2)", "PCE(%)"]
    try:
        # 3-Point Smart Start
        if context['auto_vmp']:
            logging.info("Auto-Finding Vmp via JV Scan...")
            s, e = float(context['jv_start']), float(context['jv_end'])
            _, found_vmp, _, _, _, _, _ = run_sweep_internal(context, "JV", s, e, suffix="_PreCheck")
            
            start_v = found_vmp if abs(found_vmp) > 0.01 else float(context['mppt_start'])
            v_test = [start_v - 0.02, start_v, start_v + 0.02]
            p_test = []
            
            try: keithley.query('*IDN?')
            except: rm = pyvisa.ResourceManager(); keithley = init_keithley(rm, context['gpib']); keithley.write(':OUTP ON')

            for v in v_test:
                keithley.write(f':SOUR:VOLT {v:.4f}'); time.sleep(0.2)
                i_meas, _ = measure_current_with_averaging(keithley, n_avg=3)
                p_test.append(abs(v * i_meas / float(context['area']) * 1000))
            
            start_v = v_test[np.argmax(p_test)]
            direction = 1 if p_test[2] > p_test[0] else -1
            logging.info(f"Smart Start: {start_v:.3f}V, Dir: {direction}")
            win.after(0, lambda: param_entries["MPPT Start V (V)"].delete(0, tk.END))
            win.after(0, lambda: param_entries["MPPT Start V (V)"].insert(0, f"{start_v:.3f}"))
        else:
            start_v = float(context['mppt_start']); direction = 1

        step_mv = float(context['mppt_step'])
        interval = float(context['mppt_int']); duration_h = float(context['mppt_dur'])
        area = float(context['area']); v_step = step_mv / 1000.0; duration_s = duration_h * 3600
        
        logging.info(f"MPPT Started at {start_v:.3f}V"); update_status("Tracking Pmax...", "red"); notebook.select(2)
        try: keithley.query('*IDN?') 
        except: rm = pyvisa.ResourceManager(); keithley = init_keithley(rm, context['gpib']); keithley.write(':OUTP ON')

        for k in ['SPO (PCE)', 'Power Track', 'Vmp Track', 'Imp Track']: plots_data[k]['x'] = []; plots_data[k]['y'] = []
        v_ref, p_prev = start_v, -1.0; t_start = time.time()
        
        # Initial Point Record (t=0)
        keithley.write(f':SOUR:VOLT {v_ref:.4f}'); time.sleep(0.2)
        i_init, _ = measure_current_with_averaging(keithley, n_avg=3)
        j_init = abs(i_init/area)*1000; p_init = abs(v_ref*j_init)
        data_captured.append([0.0, v_ref, i_init, p_init, p_init])
        win.after(0, update_plot_mppt, 0.0, p_init, p_init, v_ref, j_init, duration_s)
        p_prev = p_init

        while (time.time() - t_start) < duration_s and not stop_flag:
            if is_paused: time.sleep(0.5); continue
            loop_start = time.time()
            keithley.write(f':SOUR:VOLT {v_ref:.4f}'); time.sleep(0.1)
            i_meas, _ = measure_current_with_averaging(keithley, n_avg=1) # Fast read
            j_curr = abs(i_meas/area)*1000; p_curr = abs(v_ref*j_curr)
            if p_prev > 0 and p_curr < p_prev: direction *= -1
            v_ref += (direction * v_step)
            if v_ref > 1.5: v_ref=1.5; direction=-1
            if v_ref < -0.2: v_ref=-0.2; direction=1
            p_prev = p_curr
            elapsed = time.time() - t_start
            data_captured.append([round(elapsed, 1), v_ref, i_meas, p_curr, p_curr])
            if interval > 0.5: win.after(0, update_plot_mppt, elapsed, p_curr, p_curr, v_ref, j_curr, duration_s)
            else: 
                if len(data_captured) % 5 == 0: win.after(0, update_plot_mppt, elapsed, p_curr, p_curr, v_ref, j_curr, duration_s)
            wait = interval - (time.time() - loop_start)
            if wait > 0: time.sleep(wait)
    except Exception as e: logging.error(f"MPPT Error: {e}"); messagebox.showerror("Error", str(e))
    finally:
        if keithley: 
            try: 
                keithley.write(':OUTP OFF')
                keithley.close()
            except: 
                pass
        if data_captured: save_formatted_excel("MPPT", headers, data_captured, context)
        stop_exp()

def run_spo_logic(context):
    """SPO (Stabilized Power Output) - Fixed voltage tracking"""
    global keithley, stop_flag, is_paused
    data_captured = []
    headers = ["Elapsed(s)", "Voltage(V)", "Current(A)", "Power(mW/cm2)", "PCE(%)"]
    
    try:
        spo_voltage = float(get_param("SPO Voltage (V)"))
        duration_s = float(get_param("SPO Duration (s)"))
        interval = float(get_param("SPO Interval (s)"))
        area = float(context['area'])
        
        logging.info(f"SPO Started at {spo_voltage:.3f}V for {duration_s}s")
        update_status(f"SPO @ {spo_voltage:.3f}V", "darkorange")
        notebook.select(2)  # SPO (PCE) Tab
        
        try:
            keithley.query('*IDN?')
        except:
            rm = pyvisa.ResourceManager()
            keithley = init_keithley(rm, context['gpib'])
            keithley.write(':OUTP ON')
        
        # Clear previous data
        for k in ['SPO (PCE)', 'Power Track', 'Vmp Track', 'Imp Track']:
            plots_data[k]['x'] = []
            plots_data[k]['y'] = []
        
        # Set fixed voltage
        keithley.write(f':SOUR:VOLT {spo_voltage:.4f}')
        t_start = time.time()
        
        # Initial stabilization
        logging.info("Stabilizing...")
        time.sleep(2.0)
        
        while (time.time() - t_start) < duration_s and not stop_flag:
            if is_paused:
                time.sleep(0.5)
                continue
            
            loop_start = time.time()
            
            # Measure current with averaging
            i_meas, _ = measure_current_with_averaging(keithley, n_avg=3)
            j_curr = abs(i_meas / area) * 1000
            p_curr = abs(spo_voltage * j_curr)
            
            elapsed = time.time() - t_start
            data_captured.append([round(elapsed, 1), spo_voltage, i_meas, p_curr, p_curr])
            
            # Update plot
            win.after(0, update_plot_mppt, elapsed, p_curr, p_curr, spo_voltage, j_curr, duration_s)
            
            # Wait for next interval
            wait = interval - (time.time() - loop_start)
            if wait > 0:
                time.sleep(wait)
        
        # Calculate average PCE
        if data_captured:
            avg_pce = np.mean([row[4] for row in data_captured])
            std_pce = np.std([row[4] for row in data_captured])
            logging.info(f"SPO Complete: Avg PCE = {avg_pce:.2f}% ± {std_pce:.2f}%")
            
    except Exception as e:
        logging.error(f"SPO Error: {e}")
        messagebox.showerror("Error", str(e))
    finally:
        if keithley:
            try:
                keithley.write(':OUTP OFF')
                keithley.close()
            except:
                pass
        if data_captured:
            save_formatted_excel("SPO", headers, data_captured, context)
        stop_exp()

def update_plot_mppt(t, pce, power, vmp, imp, duration=0):
    plots_data['SPO (PCE)']['x'].append(t);   plots_data['SPO (PCE)']['y'].append(pce)
    plots_data['Power Track']['x'].append(t); plots_data['Power Track']['y'].append(power)
    plots_data['Vmp Track']['x'].append(t);   plots_data['Vmp Track']['y'].append(vmp)
    plots_data['Imp Track']['x'].append(t);   plots_data['Imp Track']['y'].append(imp)
    unit = param_entries["MPPT Unit"].get()
    divisors = {"sec": 1, "min": 60, "hour": 3600, "day": 86400}
    div = divisors.get(unit, 1)
    
    for key in ['SPO (PCE)', 'Power Track', 'Vmp Track', 'Imp Track']:
        p_data = plots_data[key]; ax = p_data['ax']; ax.cla()
        x_raw = np.array(p_data['x'])
        y_raw = np.array(p_data['y'])
        
        if unit == "sec":
            ax.plot(x_raw, y_raw, color=p_data['color'], lw=1.5)
            ax.set_xlabel(f"{p_data['ax_label'][0]}")
        else:
            x_scaled = x_raw / div
            bins = np.floor(x_scaled).astype(int)
            unique_bins = np.unique(bins)
            x_binned, y_mean, y_err = [], [], []
            
            for b in unique_bins:
                mask = bins == b
                if np.any(mask):
                    x_binned.append(b)
                    y_vals = y_raw[mask]
                    y_mean.append(np.mean(y_vals))
                    y_err.append(np.std(y_vals))
            
            ax.errorbar(x_binned, y_mean, yerr=y_err, fmt='o', color=p_data['color'], capsize=3, markersize=4)
            ax.set_xlabel(f"Time ({unit})")
        ax.set_ylabel(p_data['ax_label'][1])
        prog_str = f" [{t:.1f} / {duration:.1f} s]" if duration > 0 else ""
        ax.set_title(key + prog_str)
        ax.grid(True)
        p_data['canvas'].draw()

def run_loop_logic(context, loop_type):
    try: count = int(get_param("Repeat Count")); interval = float(get_param("Loop Int (s)"))
    except: logging.error("Invalid Loop Params"); return
    logging.info(f"Starting {loop_type} Loop: {count} times")
    for i in range(count):
        if stop_flag: break
        logging.info(f"Loop {i+1}/{count}..."); update_status(f"{loop_type} {i+1}/{count}", "darkblue")
        loop_suffix = f"_Loop{i+1}"
        if loop_type == "Loop_JV":
            s, e = float(context['jv_start']), float(context['jv_end'])
            run_sweep_internal(context, "JV", s, e, suffix=loop_suffix)
        elif loop_type == "Loop_QSS":
            s, e = float(context['qss_start']), float(context['qss_end'])
            run_sweep_internal(context, "QSS", s, e, suffix=loop_suffix)
        elif loop_type == "Loop_Hysteresis":
            run_hyst_logic(context) 
        if i < count - 1: time.sleep(interval)
    logging.info("Loop Finished."); stop_exp()

def start_thread(mode):
    global thread, stop_flag, is_paused
    if not param_entries["GPIB Address"].get(): messagebox.showerror("Error", "Select GPIB Address"); return
    if not global_save_path: messagebox.showerror("No Folder", "Please click 'Select Save Folder' first!"); return
    
    context = {
        'mode': mode, 'gpib': param_entries["GPIB Address"].get(),
        'user': get_param("User Name"), 'dev_id': get_param("Device ID"), 'area': get_param("Active Area (cm²)"),
        'pixel': pixel_var.get(), 'auto_vmp': auto_vmp_var.get(),
        'jv_start': get_param("JV Start V"), 'jv_end': get_param("JV End V"),
        'jv_step': get_param("JV Step (mV)"), 'jv_delay': get_param("JV Delay (ms)"),
        'qss_start': get_param("QSS Start V"), 'qss_end': get_param("QSS End V"),
        'qss_step': get_param("QSS Step (mV)"), 'qss_delay': get_param("Delay (Init)"),
        'max_wait': get_param("Max Wait (s)"), 'threshold': get_param("Threshold (%)"),
        'mppt_start': get_param("MPPT Start V (V)"), 'mppt_step': get_param("Perturb Step (mV)"),
        'mppt_int': get_param("MPPT Int (s)"), 'mppt_dur': get_param("Duration (h)"),
        'spo_voltage': get_param("SPO Voltage (V)"), 'spo_duration': get_param("SPO Duration (s)"),
        'spo_interval': get_param("SPO Interval (s)")
    }

    stop_flag, is_paused = False, False
    all_btns = [btn_jv_single, btn_qss_single, btn_mppt_single, btn_jv_loop, btn_qss_loop, btn_hyst_loop, btn_hyst_single, btn_stop, btn_qss_nrel, btn_spo]
    for btn in all_btns:
        if btn == btn_stop: btn.config(state=tk.NORMAL)
        else: btn.config(state=tk.DISABLED)
    
    if mode.startswith("Loop_"): thread = threading.Thread(target=run_loop_logic, args=(context, mode), daemon=True)
    elif mode == "Hysteresis": thread = threading.Thread(target=run_hyst_logic, args=(context,), daemon=True)
    elif mode == "QSS-NREL": thread = threading.Thread(target=run_qss_nrel_logic, args=(context,), daemon=True)
    elif mode == "SPO": thread = threading.Thread(target=run_spo_logic, args=(context,), daemon=True)
    else:
        target = run_jv_logic if mode == "JV" else (run_qss_logic if mode == "QSS" else run_mppt_logic)
        thread = threading.Thread(target=target, args=(context,), daemon=True)
    thread.start()

def stop_exp():
    global stop_flag; stop_flag = True; update_status("Stopping...", "orange")
    win.after(1000, lambda: [
        btn_jv_single.config(state=tk.NORMAL), btn_qss_single.config(state=tk.NORMAL), btn_mppt_single.config(state=tk.NORMAL),
        btn_jv_loop.config(state=tk.NORMAL), btn_qss_loop.config(state=tk.NORMAL), btn_hyst_loop.config(state=tk.NORMAL),
        btn_hyst_single.config(state=tk.NORMAL), btn_qss_nrel.config(state=tk.NORMAL), btn_spo.config(state=tk.NORMAL),
        btn_stop.config(state=tk.DISABLED),
        update_status("Ready", "gray")
    ])
def toggle_pause():
    global is_paused; is_paused = not is_paused
    btn_pause.config(text="RESUME" if is_paused else "PAUSE"); update_status("Paused" if is_paused else "Running", "orange" if is_paused else "green")

if __name__ == "__main__": setup_logger(); win.mainloop()